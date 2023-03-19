import openpyxl
import numexpr as ne
import random
from math import *

print("Введите функцию f(x), которая вас интересует:")
inputFunc = input("-->")

print("Диапазон [a, b] (через пробел, дробную часть через точку): ")
a, b = [float(i) for i in input("-->").split()]

values = [[], [], [], [], []]
n = 1
while n <= 100000:
    values[0].append(n)
    dx = round((b - a) / n, 10)

    # создание значений разбиений
    xn = []
    yn = []
    i = a
    while i <= b:
        xn.append(i)
        i = round(i + dx, 10)

    s1, s2, s3, s4 = 0, 0, 0, 0
    # ne.evaluate(inputFunc.replace("x", "(" + str(i) + ")"))
    l = ne.evaluate(inputFunc.replace("x", "(" + str(xn[0]) + ")"))
    s1 += l
    for i in range(1, len(xn)):
        r = ne.evaluate(inputFunc.replace("x", "(" + str(xn[i]) + ")"))
        if i != (len(xn)-1):
            s1 += r
        s2 += r
        s3 += ne.evaluate(inputFunc.replace("x", "(" + str(xn[i] - dx / 2) + ")"))
        s4 += ne.evaluate(inputFunc.replace("x", "(" + str(random.uniform(xn[i - 1], xn[i])) + ")"))
        l = r

    values[1].append(s1 * dx)
    values[2].append(s2 * dx)
    values[3].append(s3 * dx)
    values[4].append(s4 * dx)

    print("n = " + str(n))
    n = n * 2

# Create a new Excel workbook
workbook = openpyxl.Workbook()
# Select the first sheet
sheet = workbook.active

# Записываем заголовки столбцов
sheet.cell(row=1, column=1, value='n')
sheet.cell(row=1, column=2, value='left')
sheet.cell(row=1, column=3, value='right')
sheet.cell(row=1, column=4, value='medium')
sheet.cell(row=1, column=5, value='random')

for i in range(len(values[0])):
    row = i + 2
    sheet.cell(row=row, column=1, value=values[0][i])
    sheet.cell(row=row, column=2, value=values[1][i])
    sheet.cell(row=row, column=3, value=values[2][i])
    sheet.cell(row=row, column=4, value=values[3][i])
    sheet.cell(row=row, column=5, value=values[4][i])

# Сохраняем файл
workbook.save('Laba2.xlsx')
